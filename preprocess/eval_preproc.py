import numpy as np
import argparse
from skimage.metrics import peak_signal_noise_ratio as psnr
from skimage.metrics import structural_similarity as ssim
import nibabel as nib
import os
import openpyxl as op
from options.test_options import TestOptions


Error = {}
MSE = {}
SSIM = {}
PSNR = {}

def  get_metrics_result(resultdir, gtdir, maskdir, target_mod, ifmask):
    wb = op.Workbook()
    ws = wb.create_sheet()
    idx = 2

    patients = os.listdir(gtdir)
    patients.sort()
    for sel_patient in patients:
        patients_name = sel_patient
        print(patients_name)
        gt_nii_dir = os.path.join(gtdir, patients_name)
        for nii_file in os.listdir(gt_nii_dir):
            if target_mod in nii_file:
                gt_nii_file = nii_file
            elif "uih" in nii_file:
                uih_nii_file = nii_file
        gt = nib.load(os.path.join(gt_nii_dir, gt_nii_file)).get_fdata()
        result = nib.load(os.path.join(gt_nii_dir, uih_nii_file)).get_fdata()
        
        if not ifmask:
            result_flat = np.ravel(result)
            gt_flat = np.ravel(gt)

        else:
            mask_nii_file = gt_nii_file + '_mask.nii.gz'
            mask = nib.load(os.path.join(maskdir, patients_name, mask_nii_file)).get_fdata()
            mask_flat = np.ravel(mask)
            ind = mask_flat == 1
            result_flat = np.ravel(result)[ind]
            gt_flat = np.ravel(gt)[ind]
        
        # result_flat = result_flat/np.max(result_flat)
        # gt_flat = gt_flat/np.max(gt_flat)
        # result_flat = result_flat/255.0
        # gt_flat = gt_flat/255.0        
        error_value = np.mean(np.abs(gt_flat-result_flat))
        mse_value = np.mean(pow((gt_flat-result_flat),2))
        ssim_value = ssim(gt_flat,result_flat)
        psnr_value = psnr(gt_flat,result_flat)
        ws.cell(column=1, row=idx, value=sel_patient)
        ws.cell(column=2, row=idx, value=error_value)
        ws.cell(column=3, row=idx, value=mse_value)
        ws.cell(column=4, row=idx, value=ssim_value)
        ws.cell(column=5, row=idx, value=psnr_value)                
        idx = idx+1           
                
    ws.cell(column=2, row=1, value='Error')
    ws.cell(column=3, row=1, value='MSE')
    ws.cell(column=4, row=1, value='SSIM')
    ws.cell(column=5, row=1, value='PSNR')
    wb.save(os.path.join(resultdir, 'summary.xlsx'))
 
if __name__ == '__main__':
    resultdir = "/data1/chundanxu/pgan_npy_input/code7.3/mydata_SRPBS_aver_source_aug_3_all/results_dir/CG/"
    gtdir = "/data1/chundanxu/pgan_npy_input/code7.3/mydata_CG_aver/normalize_imgs/test/"
    maskdir = "/data1/chundanxu/pgan_npy_input/code7.3/mydata_CG_aver/mask_imgs/test/"
    get_metrics_result(resultdir=resultdir, gtdir=gtdir, maskdir=maskdir, target_mod="ge", ifmask=False)